import openpyxl
import json

def extract_rfce():
    wb = openpyxl.load_workbook('133009889-16042026193727.xlsx', data_only=True)
    if 'RFCE' not in wb.sheetnames:
        print("Sheet RFCE not found")
        return
    
    sheet = wb['RFCE']
    data = []
    # Read headers from the first row
    headers = [cell.value for cell in sheet[1]]
    
    # Read rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(row): # Skip empty rows
            data.append(dict(zip(headers, row)))
            
    with open('rfce_data.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"Extracted {len(data)} rows to rfce_data.json")

if __name__ == "__main__":
    extract_rfce()
